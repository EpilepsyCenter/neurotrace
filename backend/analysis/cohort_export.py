"""Cohort export — Phase B.8.

Three export formats, all driven from the wizard's ``design.groups``
— the same source of truth the analysis cards display, NOT the raw
folder aggregate. So whatever you see in the UI is exactly what
lands in the file:

* ``write_excel_summary``  → multi-sheet ``.xlsx``: cohort meta,
                              stats summary, cells (wide), cells
                              (long), per-event long-format
                              distributions, and ECDF tables (one
                              for the per-group means and one for
                              every individual cell).
* ``write_excel_cells``    → single-sheet ``_cells.xlsx``: one row
                              per cell with every scalar metric as
                              a column. Convenient for stats-package
                              re-import.
* ``write_prism_pzfx``     → GraphPad Prism ``.pzfx`` project,
                              built via the ``pzfx`` package so the
                              file is fully Prism-compliant. Per
                              scalar metric: a Column table with one
                              column per group. Per distribution
                              metric: pooled-events Column table,
                              group-mean ECDF XY table, and per-cell
                              ECDF XY table — exactly the shapes the
                              cohort modal uses to draw its
                              distribution graphs.

Distribution graphs match what the user sees in the cohort modal:
ECDFs (cumulative probability against the metric value) on a shared
x-grid taken from the union of all observed values. Per-cell ECDFs
+ a per-group mean curve (mean of per-cell ECDFs, NOT a re-pool —
matches the modal's no-pseudoreplication policy).
"""

from __future__ import annotations

import io
import math
import os
import re
import tempfile
from datetime import datetime
from typing import Any, Optional

import numpy as np
import pandas as pd


# ---------------------------------------------------------------------
# Group input shape — what the frontend ships in.
#
# Each group is { tag: str, cells: [cell_dict, ...] }, where each
# ``cell_dict`` is the same shape as a cell in
# ``aggregate_folder``'s output (file_path, cell_id, animal_id,
# scalars, distributions, meta, …). We trust the frontend's grouping
# completely — the wizard has already applied selectedTags,
# filterTags, seriesRole, nUnit collapse (animal/series/cell) and
# the missing-animal-id exclusion. Re-filtering server-side would
# only let drift creep in between what users see and what they save.
# ---------------------------------------------------------------------

def _flatten_groups(design_groups: list[dict]) -> list[tuple[dict, str]]:
    """``[(cell, group_tag), ...]`` in the wizard's iteration order
    so per-cell sheet rows match the on-screen card order."""
    out: list[tuple[dict, str]] = []
    for g in design_groups or []:
        tag = str(g.get('tag', ''))
        for c in (g.get('cells') or []):
            out.append((c, tag))
    return out


def _group_tags_in_order(design_groups: list[dict]) -> list[str]:
    """Group tags in iteration order. Used as the column order for
    the wide / Prism tables."""
    return [str(g.get('tag', '')) for g in (design_groups or [])]


# ---------------------------------------------------------------------
# Helpers shared across writers.
# ---------------------------------------------------------------------

def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Coerce a metric / group name into an Excel-legal sheet name.

    Excel requires:
      * length ≤ 31 chars
      * no ``: \\ / ? * [ ]`` characters
      * uniqueness within the workbook (case-insensitive)
    """
    cleaned = re.sub(r'[:\\/?*\[\]]+', '_', name)[:31] or 'sheet'
    candidate = cleaned
    n = 2
    while candidate.lower() in used:
        suffix = f'_{n}'
        candidate = (cleaned[:31 - len(suffix)] + suffix)
        n += 1
    used.add(candidate.lower())
    return candidate


def _suffixed_sheet_name(metric: str, suffix: str, used: set[str]) -> str:
    """Build a ``<metric><suffix>`` sheet name that fits Excel's 31-
    character limit while preserving the suffix verbatim. The suffix
    is the part the user reads to know what's in the sheet
    (``_groups``, ``_cells``, ``_ecdf_groups``, …) — losing it to a
    naive truncation would leave the user with ``..._timeserie_2``
    style names that don't tell you what you're looking at.

    We truncate the metric prefix instead, preserving the suffix and
    falling through to ``_safe_sheet_name``'s uniqueness handling
    for any leftover collisions (rare — only if two metrics share
    their first 31-len(suffix) characters)."""
    max_metric = max(1, 31 - len(suffix))
    return _safe_sheet_name(metric[:max_metric] + suffix, used)


def _scalar_metric_names(cells: list[dict]) -> list[str]:
    """Union of every scalar metric name observed across the
    in-export cells, in first-occurrence order."""
    seen: dict[str, None] = {}
    for c in cells:
        for k in (c.get('scalars') or {}).keys():
            if k not in seen:
                seen[k] = None
    return list(seen.keys())


def _distribution_metric_names(cells: list[dict]) -> list[str]:
    """Same as ``_scalar_metric_names`` but for distribution arrays.
    Skips empty arrays so cells with no distribution data don't
    produce empty sheets."""
    seen: dict[str, None] = {}
    for c in cells:
        for k, v in (c.get('distributions') or {}).items():
            if v and k not in seen:
                seen[k] = None
    return list(seen.keys())


def _is_timeseries_metric(cells: list[dict], metric: str) -> bool:
    """Detect whether a distribution-kind metric is actually a
    bin-by-bin time series (LTP normalised response trace,
    resistance trace, etc.) rather than a sample distribution
    (events amplitudes, IEIs, etc.).

    Cells tag this in ``meta.distribution_kinds[metric]`` — any
    cell that says ``'timeseries'`` for this metric flips the
    whole export to the timeseries layout (bins × group means)
    instead of the ECDF layout (data values × cumulative
    probability). One vote is enough: if even a single cell
    declares the metric as timeseries we treat it that way for the
    whole cohort, since mixing the two within one metric would be
    a data-shape error not a styling choice."""
    for c in cells:
        kinds = (c.get('meta') or {}).get('distribution_kinds') or {}
        if isinstance(kinds, dict) and kinds.get(metric) == 'timeseries':
            return True
    return False


def _consistent_bin_width_s(cells: list[dict]) -> Optional[float]:
    """If every cell that has a ``bin_width_s`` reports the same
    value (within 1 % rounding), return that width — drives the
    optional ``time_min`` X column. Otherwise None: bin widths
    differ across cells so a shared time axis would be misleading,
    user falls back to bin index."""
    widths: list[float] = []
    for c in cells:
        meta = c.get('meta') or {}
        w = meta.get('bin_width_s')
        try:
            wf = float(w) if w is not None else None
        except (TypeError, ValueError):
            wf = None
        if wf and wf > 0:
            widths.append(wf)
    if not widths:
        return None
    median = float(np.median(widths))
    if median <= 0:
        return None
    if all(abs(w - median) / median <= 0.01 for w in widths):
        return median
    return None


def _max_bin_count(cells: list[dict], metric: str) -> int:
    """Longest timeseries length across cells. The shared X axis
    runs 0..max-1 (or matching minutes); shorter cells are padded
    with NaN so the XY table is rectangular."""
    n_max = 0
    for c in cells:
        arr = (c.get('distributions') or {}).get(metric) or []
        if len(arr) > n_max:
            n_max = len(arr)
    return n_max


def _meta_fields(cells: list[dict]) -> list[str]:
    seen: dict[str, None] = {}
    for c in cells:
        for k in (c.get('meta') or {}).keys():
            if k not in seen:
                seen[k] = None
    return list(seen.keys())


def _excel_value(v: Any) -> Any:
    """Coerce a scalar to a value openpyxl is happy with. NaN/inf
    become None so the cell renders empty rather than ``#NUM!``."""
    if v is None:
        return None
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    if isinstance(v, (int, str, bool)):
        return v
    return str(v)


def _coerce_floats(values: Any) -> list[float]:
    out: list[float] = []
    for v in values or []:
        if v is None:
            continue
        try:
            f = float(v)
        except (TypeError, ValueError):
            continue
        if math.isnan(f) or math.isinf(f):
            continue
        out.append(f)
    return out


# ---------------------------------------------------------------------
# ECDF helpers — IDENTICAL semantics to ``cohort_graphs.plot_ecdf``
# so the export numbers match exactly what the user sees on screen.
#
#   * x_grid = sorted unique union of ALL observed values across
#     every cell of every group. Cap at 2000 sample points by
#     index to keep file size bounded; index sampling preserves the
#     density structure of the underlying data (more grid points
#     where data is dense, fewer in sparse tails).
#   * per_cell_ecdf(grid) = fraction of that cell's values ≤ x for
#     each x in grid. ``np.searchsorted(sorted_arr, grid, side='right')
#     / len(sorted_arr)``.
#   * per_group_mean_ecdf(grid) = mean across cells in the group of
#     the per-cell ECDFs at each grid point — NOT a pooled-events
#     re-ECDF (would be pseudoreplication: cells with more events
#     dominate). Same averaging policy as the cohort modal.
# ---------------------------------------------------------------------

ECDF_MAX_GRID_POINTS = 2000


def _compute_ecdf_grid(all_pooled: list[float]) -> np.ndarray:
    """Pick the shared x-grid for ECDF rendering. Inherits density
    structure from the data (vs a uniform linspace which wastes
    resolution on empty long tails — exact same logic as
    ``cohort_graphs.plot_ecdf``)."""
    arr = np.asarray(all_pooled, dtype=float)
    arr = arr[np.isfinite(arr)]
    if arr.size == 0:
        return np.array([0.0, 1.0])
    unique = np.unique(arr)
    if unique.size > ECDF_MAX_GRID_POINTS:
        idx = np.linspace(0, unique.size - 1, ECDF_MAX_GRID_POINTS).astype(int)
        return unique[idx]
    if unique.size == 1:
        v = float(unique[0])
        eps = max(abs(v) * 1e-3, 1e-9)
        return np.array([v - eps, v + eps])
    return unique


def _ecdf_at_grid(values: list[float], x_grid: np.ndarray) -> np.ndarray:
    """ECDF of ``values`` evaluated at every x in ``x_grid``.
    Matches ``np.searchsorted(sorted, grid, side='right') / n`` so
    the curve steps where values land — same convention as
    ``ax.step(values, ...)`` in matplotlib."""
    cleaned = _coerce_floats(values)
    if not cleaned:
        return np.zeros(len(x_grid), dtype=float)
    sorted_vals = np.sort(np.asarray(cleaned, dtype=float))
    return np.searchsorted(sorted_vals, x_grid, side='right') / len(sorted_vals)


# ---------------------------------------------------------------------
# Excel writers.
# ---------------------------------------------------------------------

def _write_summary_sheet(ws, stats: dict, selected_metrics: list[str]) -> None:
    headers = [
        'metric', 'n_per_group', 'test',
        'statistic', 'df', 'p', 'stars',
        'effect_size', 'effect_size_label',
        'normality', 'posthoc',
    ]
    ws.append(headers)
    for metric in selected_metrics:
        r = stats.get(metric) or {}
        if not r:
            continue
        n_per_group = ', '.join(
            f"{tag}={d.get('n', 0)}"
            for tag, d in (r.get('descriptives') or {}).items()
        )
        p = r.get('p')
        try:
            pf = float(p) if p is not None else None
        except (TypeError, ValueError):
            pf = None
        if pf is None or math.isnan(pf):
            stars = ''
        elif pf < 0.001: stars = '***'
        elif pf < 0.01: stars = '**'
        elif pf < 0.05: stars = '*'
        else: stars = 'ns'
        norm = r.get('normality') or {}
        failed = [t for t, d in norm.items() if d and d.get('verdict') == 'fail']
        any_known = any(d and d.get('verdict') != 'unknown' for d in norm.values())
        if not any_known:
            normality_label = ''
        elif not failed:
            normality_label = 'all normal'
        else:
            normality_label = f"not normal: {', '.join(failed)}"
        ph = r.get('posthoc') or []
        posthoc_label = '; '.join(
            f"{x.get('a','')} vs {x.get('b','')} p={x.get('p'):.3g}"
            for x in ph if x.get('p') is not None
        ) if ph else ''
        df_val: Any = r.get('df')
        if isinstance(df_val, (list, tuple)) and len(df_val) == 2:
            df_label = f"{df_val[0]},{df_val[1]}"
        elif isinstance(df_val, (int, float)):
            df_label = df_val
        else:
            df_label = ''
        ws.append([
            metric,
            n_per_group,
            r.get('test') or (r.get('error') and f"error: {r['error']}") or '',
            _excel_value(r.get('statistic')),
            df_label,
            _excel_value(pf),
            stars,
            _excel_value(r.get('effect_size')),
            r.get('effect_size_label') or '',
            normality_label,
            posthoc_label,
        ])


def _write_cells_wide_sheet(ws, paired: list[tuple[dict, str]],
                             scalar_metrics: list[str],
                             meta_fields: list[str]) -> None:
    # ``series_key`` is exported in the storage form (0-indexed
    # ``g:s`` or ``g:s:subtype``) — that's the join key the rest of
    # TRACER uses internally and matches the .tracer sidecar.
    # The HEKA / Patchmaster UIs show it 1-indexed (Group 1, Series 4
    # = ``0:3`` here); apply +1 to both numeric components if you need
    # to cross-reference against the recording in HEKA.
    headers = (
        ['group', 'file_name', 'file_path', 'cell_id', 'animal_id',
         'series_key (0-indexed)', 'series_specific_tags', 'group_tags']
        + list(scalar_metrics)
        + [f'meta.{k}' for k in meta_fields]
    )
    ws.append(headers)
    for c, group in paired:
        scalars = c.get('scalars') or {}
        meta = c.get('meta') or {}
        ws.append(
            [
                group,
                c.get('file_name', ''),
                c.get('file_path', ''),
                c.get('cell_id', '') or '',
                c.get('animal_id', '') or '',
                c.get('series_key', '') or '',
                ', '.join(c.get('series_specific_tags') or []),
                ', '.join(c.get('group_tags') or []),
            ]
            + [_excel_value(scalars.get(m)) for m in scalar_metrics]
            + [_excel_value(meta.get(k)) for k in meta_fields]
        )


def _write_cells_long_sheet(ws, paired: list[tuple[dict, str]],
                             scalar_metrics: list[str]) -> None:
    headers = ['cell_id', 'animal_id', 'group', 'file_name', 'metric', 'value']
    ws.append(headers)
    for c, group in paired:
        scalars = c.get('scalars') or {}
        for m in scalar_metrics:
            v = scalars.get(m)
            if v is None:
                continue
            ws.append([
                c.get('cell_id', '') or '',
                c.get('animal_id', '') or '',
                group,
                c.get('file_name', ''),
                m,
                _excel_value(v),
            ])


def _write_distribution_long_sheet(ws, paired: list[tuple[dict, str]],
                                    metric: str) -> None:
    """Per-event long format. Lets users re-bin / re-aggregate
    however they want without losing raw values."""
    headers = ['cell_id', 'animal_id', 'group', 'file_name',
               'event_index', 'value']
    ws.append(headers)
    for c, group in paired:
        arr = (c.get('distributions') or {}).get(metric) or []
        for i, v in enumerate(arr):
            try:
                fv = float(v)
                if math.isnan(fv) or math.isinf(fv):
                    continue
            except (TypeError, ValueError):
                continue
            ws.append([
                c.get('cell_id', '') or '',
                c.get('animal_id', '') or '',
                group,
                c.get('file_name', ''),
                i,
                fv,
            ])


def _write_distribution_ecdf_groups_sheet(ws,
                                            paired: list[tuple[dict, str]],
                                            ordered_groups: list[str],
                                            metric: str) -> tuple[np.ndarray, dict[str, np.ndarray]]:
    """Per-group mean ECDF on the shared x-grid. Sheet layout:
    ``value`` (the data axis) + one column per group with the
    mean cumulative probability at that x. Returns the x-grid and
    per-group mean ECDFs so the per-cell sheet shares the same
    grid (one less recompute, identical X across both sheets).
    """
    all_events: list[float] = []
    for c, _g in paired:
        all_events.extend(_coerce_floats((c.get('distributions') or {}).get(metric)))
    x_grid = _compute_ecdf_grid(all_events)

    per_group_ecdfs: dict[str, list[np.ndarray]] = {g: [] for g in ordered_groups}
    for c, g in paired:
        if g not in per_group_ecdfs:
            continue
        per_group_ecdfs[g].append(_ecdf_at_grid(
            (c.get('distributions') or {}).get(metric) or [], x_grid))

    group_means: dict[str, np.ndarray] = {}
    for g, eclist in per_group_ecdfs.items():
        if not eclist:
            group_means[g] = np.zeros(len(x_grid))
            continue
        group_means[g] = np.mean(np.stack(eclist, axis=0), axis=0)

    ws.append(['value'] + list(ordered_groups))
    for i, x in enumerate(x_grid):
        row = [float(x)]
        for g in ordered_groups:
            row.append(float(group_means[g][i]))
        ws.append(row)
    return x_grid, group_means


def _write_distribution_ecdf_cells_sheet(ws,
                                           paired: list[tuple[dict, str]],
                                           x_grid: np.ndarray,
                                           metric: str) -> None:
    """Per-cell ECDF on the same x-grid as the group-mean sheet.
    Column header per cell is ``<group>__<cell_id>`` so the user
    can immediately tell which curve belongs to which group when
    plotting in Prism / Excel."""
    headers = ['value']
    cell_columns: list[np.ndarray] = []
    for c, g in paired:
        cid = c.get('cell_id') or c.get('file_name') or '?'
        headers.append(f'{g}__{cid}')
        cell_columns.append(_ecdf_at_grid(
            (c.get('distributions') or {}).get(metric) or [], x_grid))
    ws.append(headers)
    for i, x in enumerate(x_grid):
        ws.append([float(x)] + [float(col[i]) for col in cell_columns])


def _build_timeseries_group_arrays(paired: list[tuple[dict, str]],
                                    ordered_groups: list[str],
                                    metric: str) -> tuple[
                                        int, dict[str, np.ndarray], dict[str, np.ndarray],
                                        dict[str, np.ndarray], dict[str, list[float]],
                                        dict[str, list[str]]]:
    """Per-group timeseries reductions. Returns:
      * ``n_bins``      — longest cell trace
      * ``means[g]``    — mean across cells at each bin (NaN where
                          no cell contributes)
      * ``sems[g]``     — SEM at each bin
      * ``counts[g]``   — number of cells contributing at each bin
                          (used by readers to filter sparse tails)
      * ``cells_by_group[g]`` — list of per-cell traces (for the
                                cell-level XY sheet)
      * ``cell_ids_by_group[g]`` — corresponding cell IDs

    Mean / SEM use ``nanmean`` / ``nanstd`` across ragged cell
    arrays so cells with shorter traces only contribute up to
    their length — same convention as the cohort modal's
    line+band plot.
    """
    n_bins = _max_bin_count([c for c, _g in paired], metric)
    means: dict[str, np.ndarray] = {}
    sems: dict[str, np.ndarray] = {}
    counts: dict[str, np.ndarray] = {}
    cells_by_group: dict[str, list[float]] = {g: [] for g in ordered_groups}
    cell_ids_by_group: dict[str, list[str]] = {g: [] for g in ordered_groups}

    per_group_stack: dict[str, list[np.ndarray]] = {g: [] for g in ordered_groups}
    for c, g in paired:
        if g not in per_group_stack:
            continue
        arr = (c.get('distributions') or {}).get(metric) or []
        cleaned = _coerce_floats(arr)
        if not cleaned:
            continue
        padded = np.full(n_bins, np.nan)
        padded[:len(cleaned)] = np.asarray(cleaned, dtype=float)
        per_group_stack[g].append(padded)
        cells_by_group[g].append(padded)
        cell_ids_by_group[g].append(
            str(c.get('cell_id') or c.get('file_name') or '?'))

    for g, stk in per_group_stack.items():
        if not stk:
            means[g] = np.full(n_bins, np.nan)
            sems[g] = np.full(n_bins, np.nan)
            counts[g] = np.zeros(n_bins, dtype=int)
            continue
        S = np.stack(stk, axis=0)
        with np.errstate(invalid='ignore'):
            cnt = np.sum(~np.isnan(S), axis=0)
            mean = np.nanmean(S, axis=0)
            std = np.nanstd(S, axis=0, ddof=1)
        sem = np.where(cnt > 1, std / np.sqrt(np.maximum(cnt, 1)), np.nan)
        means[g] = mean
        sems[g] = sem
        counts[g] = cnt

    return n_bins, means, sems, counts, cells_by_group, cell_ids_by_group


def _write_timeseries_groups_sheet(ws, paired: list[tuple[dict, str]],
                                    ordered_groups: list[str],
                                    metric: str) -> None:
    """XY-shape sheet for a timeseries metric, one row per bin.

    Columns: ``bin_index``, optional ``time_min`` (when bin widths
    are consistent across cells), then per group: ``<g>_mean``,
    ``<g>_sem``, ``<g>_n``. Matches the line+band plot users see
    in the cohort modal — they can replot in Prism / Excel directly
    by selecting the X column + any ``_mean`` column. ``_sem`` is
    available for error bars.
    """
    n_bins, means, sems, counts, _cells_by_group, _ids = \
        _build_timeseries_group_arrays(paired, ordered_groups, metric)
    bin_width = _consistent_bin_width_s([c for c, _g in paired])

    headers = ['bin_index']
    if bin_width:
        headers.append('time_min')
    for g in ordered_groups:
        headers += [f'{g}_mean', f'{g}_sem', f'{g}_n']
    ws.append(headers)

    for i in range(n_bins):
        row: list[Any] = [i]
        if bin_width:
            row.append(round(i * bin_width / 60.0, 6))
        for g in ordered_groups:
            row.append(_excel_value(float(means[g][i])))
            row.append(_excel_value(float(sems[g][i])))
            row.append(int(counts[g][i]))
        ws.append(row)


def _write_timeseries_cells_sheet(ws, paired: list[tuple[dict, str]],
                                   ordered_groups: list[str],
                                   metric: str) -> None:
    """One row per bin, one Y column per cell (group-prefixed).
    Lets users plot every individual recording's trace in Prism
    without having to filter the long format. Padded with NaN
    where a cell's trace runs shorter than the longest."""
    n_bins, _means, _sems, _counts, cells_by_group, ids_by_group = \
        _build_timeseries_group_arrays(paired, ordered_groups, metric)
    bin_width = _consistent_bin_width_s([c for c, _g in paired])

    headers = ['bin_index']
    if bin_width:
        headers.append('time_min')
    columns: list[np.ndarray] = []
    for g in ordered_groups:
        for cell_arr, cid in zip(cells_by_group[g], ids_by_group[g]):
            headers.append(f'{g}__{cid}')
            columns.append(cell_arr)
    ws.append(headers)

    for i in range(n_bins):
        row: list[Any] = [i]
        if bin_width:
            row.append(round(i * bin_width / 60.0, 6))
        for col in columns:
            row.append(_excel_value(float(col[i])))
        ws.append(row)


def _write_cohort_meta_sheet(ws, aggregate: dict, design: Optional[dict],
                              n_cells_in_export: int) -> None:
    ws.append(['Cohort export'])
    ws.append(['Generated', datetime.utcnow().isoformat(timespec='seconds') + 'Z'])
    ws.append(['Folder', str(aggregate.get('folder', ''))])
    ws.append(['Analysis type', str(aggregate.get('analysis_type', ''))])
    summary = aggregate.get('summary') or {}
    ws.append(['Cells in folder', summary.get('n_cells', '')])
    ws.append(['Cells included in this export', n_cells_in_export])
    ws.append(['Files scanned', summary.get('n_files_scanned', '')])
    if design:
        ws.append([])
        ws.append(['Design'])
        ws.append(['Comparison', design.get('comparison_shape', '')])
        ws.append(['Selected tags', ', '.join(design.get('selected_tags', []) or [])])
        ws.append(['Filter tags', ', '.join(design.get('filter_tags', []) or [])])
        ws.append(['Series role', design.get('series_role', '')])
        ws.append(['N unit', design.get('n_unit', '')])
        ws.append(['Test override', design.get('test_override', '')])
        # Subsampling state (B.5) — surface what the user picked so
        # the export is self-documenting.
        ss = design.get('subsample') or {}
        if ss:
            mode = ss.get('mode', 'all')
            n = ss.get('n')
            label = mode if mode == 'all' else (f"{mode} {n}" if n else f"{mode} (auto)")
            ws.append(['Events per cell', label])


def write_excel_summary(aggregate: dict,
                        stats: Optional[dict] = None,
                        selected_metrics: Optional[list[str]] = None,
                        design: Optional[dict] = None,
                        design_groups: Optional[list[dict]] = None) -> bytes:
    """Multi-sheet workbook export using the wizard's
    ``design_groups`` as the source of truth for which cells appear
    in each group. When ``design_groups`` is None we fall back to
    treating every cell in the aggregate as one anonymous group —
    useful when there's no design yet (e.g. user just aggregated
    and wants a quick raw dump)."""
    from openpyxl import Workbook  # lazy

    if design_groups:
        paired = _flatten_groups(design_groups)
        ordered_groups = _group_tags_in_order(design_groups)
    else:
        paired = [(c, '') for c in (aggregate.get('cells') or [])]
        ordered_groups = ['']
    cells = [c for c, _g in paired]

    scalar_metrics = _scalar_metric_names(cells)
    distribution_metrics = _distribution_metric_names(cells)
    meta_fields = _meta_fields(cells)
    sel = list(selected_metrics or scalar_metrics)

    wb = Workbook()
    used_names: set[str] = set()

    ws = wb.active
    ws.title = _safe_sheet_name('Cohort', used_names)
    _write_cohort_meta_sheet(ws, aggregate, design, len(paired))

    if stats:
        ws = wb.create_sheet(_safe_sheet_name('Stats summary', used_names))
        _write_summary_sheet(ws, stats, sel)

    ws = wb.create_sheet(_safe_sheet_name('Cells (wide)', used_names))
    _write_cells_wide_sheet(ws, paired, scalar_metrics, meta_fields)

    ws = wb.create_sheet(_safe_sheet_name('Cells (long)', used_names))
    _write_cells_long_sheet(ws, paired, scalar_metrics)

    # Per-metric distribution sheets. Branch on kind:
    #   * timeseries (LTP normalised response, resistance traces…):
    #     one wide XY sheet with bin × group means/SEM, one wide XY
    #     sheet with bin × per-cell traces. ECDF / per-event long
    #     format make no sense here (each "value" is a bin not an
    #     event), so we skip them.
    #   * samples (event amplitudes, IEIs, AUCs…): the existing
    #     three-sheet layout — long-format events, per-group mean
    #     ECDF, per-cell ECDF.
    for m in distribution_metrics:
        if _is_timeseries_metric(cells, m):
            ws_g = wb.create_sheet(_suffixed_sheet_name(m, '_groups', used_names))
            _write_timeseries_groups_sheet(ws_g, paired, ordered_groups, m)
            ws_c = wb.create_sheet(_suffixed_sheet_name(m, '_cells', used_names))
            _write_timeseries_cells_sheet(ws_c, paired, ordered_groups, m)
        else:
            ws_a = wb.create_sheet(_suffixed_sheet_name(m, '_events', used_names))
            _write_distribution_long_sheet(ws_a, paired, m)
            ws_b = wb.create_sheet(_suffixed_sheet_name(m, '_ecdf_groups', used_names))
            x_grid, _gm = _write_distribution_ecdf_groups_sheet(
                ws_b, paired, ordered_groups, m)
            ws_c = wb.create_sheet(_suffixed_sheet_name(m, '_ecdf_cells', used_names))
            _write_distribution_ecdf_cells_sheet(ws_c, paired, x_grid, m)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_excel_cells(aggregate: dict,
                       design: Optional[dict] = None,
                       design_groups: Optional[list[dict]] = None) -> bytes:
    """Lean per-cell export — single sheet, one row per cell, every
    scalar metric as a column."""
    from openpyxl import Workbook

    if design_groups:
        paired = _flatten_groups(design_groups)
    else:
        paired = [(c, '') for c in (aggregate.get('cells') or [])]
    cells = [c for c, _g in paired]
    scalar_metrics = _scalar_metric_names(cells)
    meta_fields = _meta_fields(cells)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Cells'
    _write_cells_wide_sheet(ws, paired, scalar_metrics, meta_fields)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------
# Prism .pzfx writer — uses the ``pzfx`` package so output is fully
# Prism-compliant XML. Tables are built from pandas DataFrames; XY
# tables identify the X column via ``x_col`` so Prism opens them
# with the data axis on X (matching the cohort modal).
# ---------------------------------------------------------------------

def _df_columns_per_group(paired: list[tuple[dict, str]],
                           ordered_groups: list[str],
                           metric: str,
                           is_distribution: bool) -> pd.DataFrame:
    """Wide DataFrame: one column per group. Scalar → each cell
    contributes one row; distribution → all events from all cells
    of that group, pooled. Padded with NaN so columns are
    rectangular (Prism reads NaN as empty cell)."""
    by_group: dict[str, list[float]] = {g: [] for g in ordered_groups}
    for c, g in paired:
        if g not in by_group:
            continue
        if is_distribution:
            by_group[g].extend(_coerce_floats(
                (c.get('distributions') or {}).get(metric)))
        else:
            v = (c.get('scalars') or {}).get(metric)
            if v is None:
                continue
            try:
                by_group[g].append(float(v))
            except (TypeError, ValueError):
                continue
    if not any(by_group.values()):
        return pd.DataFrame(columns=ordered_groups)
    n_max = max(len(vs) for vs in by_group.values())
    padded = {
        g: vs + [float('nan')] * (n_max - len(vs))
        for g, vs in by_group.items()
    }
    return pd.DataFrame(padded, columns=ordered_groups)


def _df_ecdf_groups(paired: list[tuple[dict, str]],
                     ordered_groups: list[str],
                     metric: str) -> tuple[pd.DataFrame, np.ndarray]:
    """XY DataFrame: ``value`` (X) + one column per group holding
    that group's mean ECDF at each x. Returns the x-grid so the
    per-cell DataFrame shares the same X axis."""
    all_events: list[float] = []
    for c, _g in paired:
        all_events.extend(_coerce_floats(
            (c.get('distributions') or {}).get(metric)))
    x_grid = _compute_ecdf_grid(all_events)

    per_group: dict[str, list[np.ndarray]] = {g: [] for g in ordered_groups}
    for c, g in paired:
        if g not in per_group:
            continue
        per_group[g].append(_ecdf_at_grid(
            (c.get('distributions') or {}).get(metric) or [], x_grid))

    cols: dict[str, np.ndarray] = {'value': x_grid}
    for g in ordered_groups:
        if per_group[g]:
            cols[g] = np.mean(np.stack(per_group[g], axis=0), axis=0)
        else:
            cols[g] = np.zeros(len(x_grid))
    return pd.DataFrame(cols), x_grid


def _df_timeseries_groups(paired: list[tuple[dict, str]],
                           ordered_groups: list[str],
                           metric: str) -> pd.DataFrame:
    """XY DataFrame for a timeseries metric: ``bin_index`` (X),
    optional ``time_min`` column (when bin widths consistent),
    then per-group mean/SEM/n columns. Same numbers as the cohort
    modal's line + band plot."""
    n_bins, means, sems, counts, _cbg, _ibg = _build_timeseries_group_arrays(
        paired, ordered_groups, metric)
    bin_width = _consistent_bin_width_s([c for c, _g in paired])

    cols: dict[str, np.ndarray] = {'bin_index': np.arange(n_bins, dtype=int)}
    if bin_width:
        cols['time_min'] = np.arange(n_bins) * (bin_width / 60.0)
    for g in ordered_groups:
        cols[f'{g}_mean'] = means[g]
        cols[f'{g}_sem'] = sems[g]
        cols[f'{g}_n'] = counts[g].astype(int)
    return pd.DataFrame(cols)


def _df_timeseries_cells(paired: list[tuple[dict, str]],
                          ordered_groups: list[str],
                          metric: str) -> pd.DataFrame:
    """XY DataFrame for a timeseries metric: ``bin_index`` (X),
    optional ``time_min``, then one Y column per cell named
    ``<group>__<cell_id>``."""
    n_bins, _means, _sems, _counts, cells_by_group, ids_by_group = \
        _build_timeseries_group_arrays(paired, ordered_groups, metric)
    bin_width = _consistent_bin_width_s([c for c, _g in paired])

    cols: dict[str, np.ndarray] = {'bin_index': np.arange(n_bins, dtype=int)}
    if bin_width:
        cols['time_min'] = np.arange(n_bins) * (bin_width / 60.0)
    used: set[str] = set(cols.keys())
    for g in ordered_groups:
        for cell_arr, cid in zip(cells_by_group[g], ids_by_group[g]):
            base = f'{g}__{cid}'
            unique = base
            suffix = 2
            while unique in used:
                unique = f'{base}_{suffix}'
                suffix += 1
            used.add(unique)
            cols[unique] = cell_arr
    return pd.DataFrame(cols)


def _df_ecdf_cells(paired: list[tuple[dict, str]],
                    x_grid: np.ndarray,
                    metric: str) -> pd.DataFrame:
    """XY DataFrame: ``value`` (X) + one Y column per cell, each
    column named ``<group>__<cell_id>``."""
    cols: dict[str, np.ndarray] = {'value': x_grid}
    for c, g in paired:
        cid = c.get('cell_id') or c.get('file_name') or '?'
        col_name = f'{g}__{cid}'
        unique = col_name
        suffix = 2
        while unique in cols:
            unique = f'{col_name}_{suffix}'
            suffix += 1
        cols[unique] = _ecdf_at_grid(
            (c.get('distributions') or {}).get(metric) or [], x_grid)
    return pd.DataFrame(cols)


def write_prism_pzfx(aggregate: dict,
                      design: Optional[dict] = None,
                      design_groups: Optional[list[dict]] = None) -> bytes:
    """Emit a fully Prism-compliant ``.pzfx`` project via the
    ``pzfx`` package. Cells come from the wizard's ``design_groups``;
    group columns use those exact tag strings as Prism column titles.

    Per scalar metric → one Column table (one column per group).
    Per distribution metric → three tables: pooled events (Column),
    group-mean ECDF (XY, X = value), per-cell ECDF (XY, X = value).
    """
    from pzfx import write_pzfx  # lazy

    if design_groups:
        paired = _flatten_groups(design_groups)
        ordered_groups = _group_tags_in_order(design_groups)
    else:
        paired = [(c, '') for c in (aggregate.get('cells') or [])]
        ordered_groups = ['']
    cells = [c for c, _g in paired]

    scalar_metrics = _scalar_metric_names(cells)
    distribution_metrics = _distribution_metric_names(cells)

    tables: dict[str, pd.DataFrame] = {}
    x_col_map: dict[str, str] = {}  # XY-only

    for m in scalar_metrics:
        df = _df_columns_per_group(paired, ordered_groups, m, is_distribution=False)
        if not df.empty:
            tables[m] = df

    for m in distribution_metrics:
        if _is_timeseries_metric(cells, m):
            # Timeseries: bin × group means/SEM and bin × per-cell.
            # X column is ``bin_index``; ``time_min`` (if present)
            # rides along as an additional non-X numeric column —
            # users can swap to it as Prism's X axis after import.
            df_g = _df_timeseries_groups(paired, ordered_groups, m)
            tables[f'{m}_groups'] = df_g
            x_col_map[f'{m}_groups'] = 'bin_index'
            df_c = _df_timeseries_cells(paired, ordered_groups, m)
            tables[f'{m}_cells'] = df_c
            x_col_map[f'{m}_cells'] = 'bin_index'
        else:
            df_events = _df_columns_per_group(paired, ordered_groups, m, is_distribution=True)
            if not df_events.empty:
                tables[f'{m}_events'] = df_events
            df_eg, x_grid = _df_ecdf_groups(paired, ordered_groups, m)
            tables[f'{m}_ecdf_groups'] = df_eg
            x_col_map[f'{m}_ecdf_groups'] = 'value'
            df_ec = _df_ecdf_cells(paired, x_grid, m)
            tables[f'{m}_ecdf_cells'] = df_ec
            x_col_map[f'{m}_ecdf_cells'] = 'value'

    # ``pzfx.write_pzfx`` accepts a per-table x_col list aligned
    # with iteration order. None means "no X column for this table".
    x_col_list: list[Optional[str]] = []
    for name in tables.keys():
        x_col_list.append(x_col_map.get(name))

    with tempfile.NamedTemporaryFile(suffix='.pzfx', delete=False) as tmp:
        tmp_path = tmp.name
    try:
        write_pzfx(
            tables, tmp_path,
            row_names=False,
            x_col=x_col_list if any(x for x in x_col_list) else None,
            n_digits=6,
        )
        with open(tmp_path, 'rb') as f:
            return f.read()
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
